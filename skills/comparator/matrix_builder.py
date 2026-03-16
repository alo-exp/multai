#!/usr/bin/env python3
"""
Build a comparison matrix XLSX from scratch.

Produces a fully styled XLSX with:
  - Row 1: Title (merged across all columns)
  - Row 2: Column headers (Capability / Feature | Priority | Platform1 | ...)
  - Row 3: COUNTIF row (total capabilities per platform)
  - Row 4: Score row (priority-weighted COUNTIFS per platform)
  - Row 5+: Data rows (category headings + feature rows with ticks)
  - Freeze panes at A5
  - Auto-filter on header row

Input: a JSON config file:
{
  "title": "Capabilities Comparison Matrix",
  "categories": [
    {
      "name": "1. Category Name",
      "features": [
        {"name": "Feature A", "priority": "High"},
        {"name": "Feature B", "priority": "Critical"}
      ]
    }
  ],
  "platforms": [
    {"name": "Platform1", "features": ["Feature A", "Feature B"]},
    {"name": "Platform2", "features": ["Feature A"]}
  ]
}

When an existing matrix XLSX is provided via --clone-style, styles are
cloned from it. Otherwise, built-in defaults are used.

CLI:
    python3 matrix_builder.py --config build.json --out matrix.xlsx [--clone-style existing.xlsx]
"""
from __future__ import annotations

import argparse
import copy
import json
import logging
import sys
from pathlib import Path
from typing import Any, Optional

import openpyxl
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

log = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Layout constants (new matrices always use the "with-title" layout)
# ---------------------------------------------------------------------------
TITLE_ROW = 1
HEADER_ROW = 2
TOTAL_ROW = 3
SCORE_ROW = 4
DATA_START = 5
FEAT_COL = 1
PRIO_COL = 2
PLAT_START = 3

TICK = "\u2714"

WEIGHTS: dict[str, int] = {
    "Critical": 5,
    "Very High": 4,
    "High": 3,
    "Medium": 2,
    "Low": 1,
}

# ---------------------------------------------------------------------------
# Default styles (used when no --clone-style is provided)
# ---------------------------------------------------------------------------
_THIN = Side(style="thin")
_DEFAULT_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_TITLE_FONT = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
_TITLE_FILL = PatternFill("solid", fgColor="2F5496")
_TITLE_ALIGN = Alignment(horizontal="center", vertical="center")

_HEADER_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
_HEADER_FILL = PatternFill("solid", fgColor="4472C4")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

_TOTAL_FONT = Font(name="Calibri", bold=True, size=11)
_TOTAL_FILL = PatternFill("solid", fgColor="D6E4F0")
_TOTAL_ALIGN = Alignment(horizontal="center", vertical="center")

_SCORE_FONT = Font(name="Calibri", bold=True, size=11)
_SCORE_FILL = PatternFill("solid", fgColor="D6E4F0")
_SCORE_ALIGN = Alignment(horizontal="center", vertical="center")

_CAT_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
_CAT_FILL = PatternFill("solid", fgColor="5B9BD5")
_CAT_ALIGN = Alignment(horizontal="left", vertical="center")

_FEAT_FONT = Font(name="Calibri", size=10)
_FEAT_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)

_TICK_FONT = Font(name="Calibri", size=10, color="006100")
_TICK_FILL = PatternFill("solid", fgColor="C6EFCE")
_TICK_ALIGN = Alignment(horizontal="center", vertical="center")

_EMPTY_FILL = PatternFill("solid", fgColor="FFFFFF")

PRIORITY_FILLS: dict[str, PatternFill] = {
    "Critical": PatternFill("solid", fgColor="FF9999"),
    "Very High": PatternFill("solid", fgColor="FFB366"),
    "High": PatternFill("solid", fgColor="FFDD57"),
    "Medium": PatternFill("solid", fgColor="D6E4F0"),
    "Low": PatternFill("solid", fgColor="E2EFDA"),
}

_PRIO_FONT = Font(name="Calibri", size=10)
_PRIO_ALIGN = Alignment(horizontal="center", vertical="center")


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _countif(col_letter: str) -> str:
    return f'=COUNTIF({col_letter}{DATA_START}:{col_letter}1048576,"?*")'


def _score_formula(col_letter: str) -> str:
    b = f"$B${DATA_START}:$B$1048576"
    c = f"{col_letter}${DATA_START}:{col_letter}$1048576"
    parts = []
    for prio, wt in WEIGHTS.items():
        parts.append(f'COUNTIFS({b},"{prio}",{c},"?*")*{wt}')
    return "=" + "+".join(parts)


def _style_cell(cell, font=None, fill=None, align=None, border=None):
    """Apply style to a cell (only non-None attributes)."""
    if font:
        cell.font = copy.copy(font)
    if fill:
        cell.fill = copy.copy(fill)
    if align:
        cell.alignment = copy.copy(align)
    if border:
        cell.border = copy.copy(border)


# ---------------------------------------------------------------------------
# Builder
# ---------------------------------------------------------------------------

def build_matrix(config: dict, out_xlsx: str, clone_xlsx: Optional[str] = None) -> dict:
    """Build a comparison matrix from JSON config.

    Args:
        config: {title, categories: [{name, features: [{name, priority}]}],
                 platforms: [{name, features: [feat_name, ...]}]}
        out_xlsx: Output path.
        clone_xlsx: Optional existing XLSX to clone styles from.

    Returns:
        {platforms_added, categories, features, total_rows}
    """
    if clone_xlsx is not None:
        log.warning(
            "--clone-style / clone_xlsx is not yet implemented; "
            "the new matrix will use default openpyxl styles."
        )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Comparison Matrix"

    title = config.get("title", "Capabilities Comparison Matrix")
    categories = config.get("categories", [])
    platforms = config.get("platforms", [])

    num_plats = len(platforms)
    last_col = PLAT_START + num_plats - 1
    last_ltr = get_column_letter(last_col)

    # ------------------------------------------------------------------
    # Row 1: Title
    # ------------------------------------------------------------------
    ws.merge_cells(f"A{TITLE_ROW}:{last_ltr}{TITLE_ROW}")
    title_cell = ws.cell(TITLE_ROW, 1)
    title_cell.value = title
    _style_cell(title_cell, _TITLE_FONT, _TITLE_FILL, _TITLE_ALIGN, _DEFAULT_BORDER)
    ws.row_dimensions[TITLE_ROW].height = 30

    # ------------------------------------------------------------------
    # Row 2: Headers
    # ------------------------------------------------------------------
    h_cap = ws.cell(HEADER_ROW, FEAT_COL)
    h_cap.value = "Capability / Feature"
    _style_cell(h_cap, _HEADER_FONT, _HEADER_FILL, _HEADER_ALIGN, _DEFAULT_BORDER)

    h_prio = ws.cell(HEADER_ROW, PRIO_COL)
    h_prio.value = "Priority"
    _style_cell(h_prio, _HEADER_FONT, _HEADER_FILL, _HEADER_ALIGN, _DEFAULT_BORDER)

    for p_idx, plat in enumerate(platforms):
        col = PLAT_START + p_idx
        cell = ws.cell(HEADER_ROW, col)
        cell.value = plat["name"]
        _style_cell(cell, _HEADER_FONT, _HEADER_FILL, _HEADER_ALIGN, _DEFAULT_BORDER)

    ws.row_dimensions[HEADER_ROW].height = 30

    # ------------------------------------------------------------------
    # Row 3: COUNTIF (Total Capabilities)
    # ------------------------------------------------------------------
    tc = ws.cell(TOTAL_ROW, FEAT_COL)
    tc.value = "Total Capabilities"
    _style_cell(tc, _TOTAL_FONT, _TOTAL_FILL, Alignment(horizontal="left", vertical="center"), _DEFAULT_BORDER)

    tc_p = ws.cell(TOTAL_ROW, PRIO_COL)
    tc_p.value = "\u2014"
    _style_cell(tc_p, _TOTAL_FONT, _TOTAL_FILL, _TOTAL_ALIGN, _DEFAULT_BORDER)

    for p_idx in range(num_plats):
        col = PLAT_START + p_idx
        ltr = get_column_letter(col)
        cell = ws.cell(TOTAL_ROW, col)
        cell.value = _countif(ltr)
        _style_cell(cell, _TOTAL_FONT, _TOTAL_FILL, _TOTAL_ALIGN, _DEFAULT_BORDER)

    # ------------------------------------------------------------------
    # Row 4: Score (Weighted)
    # ------------------------------------------------------------------
    sc = ws.cell(SCORE_ROW, FEAT_COL)
    sc.value = "Score"
    _style_cell(sc, _SCORE_FONT, _SCORE_FILL, Alignment(horizontal="left", vertical="center"), _DEFAULT_BORDER)

    sc_p = ws.cell(SCORE_ROW, PRIO_COL)
    sc_p.value = "\u2014"
    _style_cell(sc_p, _SCORE_FONT, _SCORE_FILL, _SCORE_ALIGN, _DEFAULT_BORDER)

    for p_idx in range(num_plats):
        col = PLAT_START + p_idx
        ltr = get_column_letter(col)
        cell = ws.cell(SCORE_ROW, col)
        cell.value = _score_formula(ltr)
        _style_cell(cell, _SCORE_FONT, _SCORE_FILL, _SCORE_ALIGN, _DEFAULT_BORDER)

    # ------------------------------------------------------------------
    # Row 5+: Data (categories + features)
    # ------------------------------------------------------------------
    # Build per-platform feature sets for fast lookup
    plat_feats: list[set[str]] = []
    for plat in platforms:
        plat_feats.append(set(plat.get("features", [])))

    current_row = DATA_START
    total_features = 0

    for cat in categories:
        # Category heading row (merged)
        ws.merge_cells(f"A{current_row}:{last_ltr}{current_row}")
        cat_cell = ws.cell(current_row, FEAT_COL)
        cat_cell.value = cat["name"]
        _style_cell(cat_cell, _CAT_FONT, _CAT_FILL, _CAT_ALIGN, _DEFAULT_BORDER)
        ws.row_dimensions[current_row].height = 22
        current_row += 1

        for feat in cat.get("features", []):
            feat_name = feat["name"]
            prio = feat.get("priority", "Medium")
            total_features += 1

            # Col A: feature name
            fc = ws.cell(current_row, FEAT_COL)
            fc.value = feat_name
            _style_cell(fc, _FEAT_FONT, _EMPTY_FILL, _FEAT_ALIGN, _DEFAULT_BORDER)

            # Col B: priority
            pc = ws.cell(current_row, PRIO_COL)
            pc.value = prio
            prio_fill = PRIORITY_FILLS.get(prio, _EMPTY_FILL)
            _style_cell(pc, _PRIO_FONT, prio_fill, _PRIO_ALIGN, _DEFAULT_BORDER)

            # Platform columns: tick or empty
            for p_idx in range(num_plats):
                col = PLAT_START + p_idx
                cell = ws.cell(current_row, col)
                cell.border = copy.copy(_DEFAULT_BORDER)
                if feat_name in plat_feats[p_idx]:
                    cell.value = TICK
                    _style_cell(cell, _TICK_FONT, _TICK_FILL, _TICK_ALIGN, _DEFAULT_BORDER)
                else:
                    cell.value = None
                    _style_cell(cell, _FEAT_FONT, _EMPTY_FILL, _TICK_ALIGN, _DEFAULT_BORDER)

            current_row += 1

    # ------------------------------------------------------------------
    # Column widths
    # ------------------------------------------------------------------
    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["B"].width = 12
    for p_idx in range(num_plats):
        ltr = get_column_letter(PLAT_START + p_idx)
        ws.column_dimensions[ltr].width = 18

    # ------------------------------------------------------------------
    # Freeze panes & auto-filter
    # ------------------------------------------------------------------
    ws.freeze_panes = f"A{DATA_START}"
    ws.auto_filter.ref = f"A{HEADER_ROW}:{last_ltr}{HEADER_ROW}"

    wb.save(out_xlsx)

    return {
        "platforms_added": num_plats,
        "platform_names": [p["name"] for p in platforms],
        "categories": len(categories),
        "features": total_features,
        "total_rows": current_row - 1,
        "output": out_xlsx,
    }


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(
        description="Build a comparison matrix XLSX from JSON config",
    )
    parser.add_argument("--config", required=True,
                        help="JSON config file with title, categories, platforms")
    parser.add_argument("--out", required=True,
                        help="Output XLSX path")
    parser.add_argument("--clone-style", default=None,
                        help="Optional existing XLSX to clone styles from (reserved)")
    args = parser.parse_args()

    config = json.loads(Path(args.config).read_text(encoding="utf-8"))
    result = build_matrix(config, args.out, clone_xlsx=args.clone_style)
    print(json.dumps(result, indent=2, ensure_ascii=False))


if __name__ == "__main__":
    main()
