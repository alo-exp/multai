"""Build Stagehand vs Playwright vs Browser-Use comparison matrix for MultAI."""

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Font, GradientFill, PatternFill, Side
)
from openpyxl.utils import get_column_letter

OUTPUT = "reports/stagehand-comparison.xlsx"

# ── Palette ──────────────────────────────────────────────────────────────────
C_HEADER_BG   = "1F4E79"   # dark navy  – column headers
C_HEADER_FG   = "FFFFFF"   # white      – column header text
C_CAT_BG      = "D6E4F0"   # light blue – category rows
C_CAT_FG      = "1F4E79"   # navy       – category text
C_ALT         = "F2F7FB"   # very light blue – alternating rows
C_WHITE       = "FFFFFF"
C_GREEN_BG    = "E2EFDA"   # light green
C_RED_BG      = "FCE4D6"   # light red/orange
C_YELLOW_BG   = "FFF2CC"   # light yellow

C_YES         = "375623"   # dark green  – full support text
C_PARTIAL     = "7D4900"   # amber       – partial support text
C_NO          = "9C0006"   # dark red    – no support text

THIN_SIDE  = Side(border_style="thin",   color="B8CCE4")
MED_SIDE   = Side(border_style="medium", color="1F4E79")

def thin_border():
    return Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)

def med_border():
    return Border(left=MED_SIDE, right=MED_SIDE, top=MED_SIDE, bottom=MED_SIDE)

def hfill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, color="000000", size=11, name="Arial"):
    return Font(name=name, bold=bold, color=color, size=size)

def centered(wrap=False):
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)

def left_align(wrap=True):
    return Alignment(horizontal="left", vertical="center", wrap_text=wrap)

def set_col_width(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width

def write_header_row(ws, row, values, bg=C_HEADER_BG, fg=C_HEADER_FG, sizes=None):
    sizes = sizes or [11] * len(values)
    for c_idx, (val, sz) in enumerate(zip(values, sizes), 1):
        cell = ws.cell(row=row, column=c_idx, value=val)
        cell.fill = hfill(bg)
        cell.font = font(bold=True, color=fg, size=sz)
        cell.alignment = centered(wrap=True)
        cell.border = thin_border()

def write_category(ws, row, label, ncols):
    cell = ws.cell(row=row, column=1, value=label)
    cell.fill = hfill(C_CAT_BG)
    cell.font = font(bold=True, color=C_CAT_FG, size=10)
    cell.alignment = left_align(wrap=False)
    cell.border = thin_border()
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    ws.row_dimensions[row].height = 18

def write_data_row(ws, row, values, alt=False):
    bg = C_ALT if alt else C_WHITE
    for c_idx, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=c_idx, value=val)
        cell.border = thin_border()
        cell.alignment = left_align(wrap=True) if c_idx == 1 else centered(wrap=True)
        if c_idx == 1:
            cell.font = font(bold=False, size=9, color="1F4E79")
            cell.fill = hfill("EBF3FB" if alt else "F7FBFF")
        else:
            # Colour-code by content
            v = str(val)
            if v.startswith("✅"):
                cell.font = font(color=C_YES, size=9)
                cell.fill = hfill("EAF4E2" if alt else C_GREEN_BG)
            elif v.startswith("⚠️"):
                cell.font = font(color=C_PARTIAL, size=9)
                cell.fill = hfill("FEF9EC" if alt else C_YELLOW_BG)
            elif v.startswith("❌"):
                cell.font = font(color=C_NO, size=9)
                cell.fill = hfill("FDF3F0" if alt else C_RED_BG)
            else:
                cell.font = font(size=9)
                cell.fill = hfill(bg)
    ws.row_dimensions[row].height = 32


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 1 — Capability-Features Matrix
# ══════════════════════════════════════════════════════════════════════════════

MATRIX_COLS = [
    "Feature / Capability",
    "Playwright\n(v1.58, Python)",
    "Browser-Use\n(v0.12.2, Python)",
    "Stagehand\n(v3.2, TypeScript)",
    "Stagehand\n(Python SDK)",
]

# (Feature label, Playwright, Browser-Use, Stagehand TS, Stagehand Python)
MATRIX_DATA = [
    # CATEGORY: BASICS
    ("__cat__", "1. BASICS & FUNDAMENTALS"),
    ("Primary Language", "Python", "Python", "TypeScript", "Python (HTTP client)"),
    ("Current Stable Version", "1.58.0", "0.12.2", "3.2.0", "Mirrors server v3.6.3"),
    ("License", "Apache 2.0", "MIT", "MIT", "MIT"),
    ("GitHub Stars", "~70,000", "~86,000", "~21,800", "Bundled with TS repo"),
    ("Framework Maturity", "⚠️ Production (testing-first)", "⚠️ Rapid iteration (sub-1.0)", "✅ Production (automation-first)", "⚠️ Newer, fewer features"),
    ("Python Native (asyncio)", "✅ Full async/await", "✅ Full asyncio", "❌ No", "⚠️ HTTP client only"),

    # CATEGORY: AUTOMATION APPROACH
    ("__cat__", "2. AUTOMATION APPROACH"),
    ("Control Model", "✅ Fully deterministic", "❌ Fully agentic (LLM drives all)", "✅ Hybrid (dev chooses AI vs code)", "✅ Hybrid (via server)"),
    ("Natural Language Instructions", "❌ None", "✅ Full (task string)", "✅ Full (act/observe/extract/agent)", "✅ Via server API"),
    ("Deterministic Selector Support", "✅ CSS, XPath, locators, roles", "⚠️ Numeric DOM index (fragile)", "✅ XPath + CSS + AI-resolved", "✅ Via server API"),
    ("Self-Healing Selectors", "❌ Breaks on UI change", "✅ LLM adapts to new structure", "✅ AI re-engages on cache miss", "✅ Via server API"),
    ("Action Caching / Replay", "❌ None", "⚠️ Cloud SDK only (OSS: none)", "✅ Auto-caching (AI→deterministic)", "✅ Via server API"),
    ("Per-Step AI Fallback", "❌ Not built-in", "✅ LLM handles every step", "✅ act() / observe() per step", "✅ Via server API"),

    # CATEGORY: BROWSER SUPPORT
    ("__cat__", "3. BROWSER & DRIVER SUPPORT"),
    ("Chrome / Chromium", "✅ Full", "✅ Full (via Playwright)", "✅ Full (CDP-native)", "✅ Via server"),
    ("Firefox", "✅ Full", "⚠️ Limited (Playwright-dependent)", "❌ Not supported (CDP-only)", "❌ Not supported"),
    ("WebKit / Safari", "✅ Full", "⚠️ Limited", "❌ Not supported", "❌ Not supported"),
    ("CDP Native Connection", "✅ connect_over_cdp()", "✅ Via Playwright CDP", "✅ v3 CDP-native (no Playwright)", "✅ Via server"),
    ("Attaches to Running Browser", "✅ Via CDP ws endpoint", "✅ Via BrowserSession(cdp_url)", "✅ connectURL() / CDP bridge", "✅ Via server"),
    ("Multi-Tab Support", "✅ context.new_page()", "✅ Via action primitives", "✅ context.newPage()", "✅ Via server"),

    # CATEGORY: CORE CAPABILITIES
    ("__cat__", "4. CORE BROWSER CAPABILITIES"),
    ("Navigation (goto, back, forward)", "✅ Full", "✅ Full (via LLM action)", "✅ Full", "✅ Via server"),
    ("Click / Hover / Keyboard", "✅ Full", "✅ Full (via LLM)", "✅ Full (act() + keyboard)", "✅ Via server"),
    ("Form Fill & Submit", "✅ fill(), type(), press()", "✅ Via LLM input_text()", "✅ act() + fillForm tool", "✅ Via server"),
    ("File Upload", "✅ setInputFiles()", "⚠️ Limited", "⚠️ Limited in SDK", "⚠️ Limited"),
    ("Screenshot Capture", "✅ page.screenshot()", "✅ Built-in per step", "✅ screenshot tool in agent", "✅ Via server"),
    ("Network Interception", "✅ route(), on('request')", "❌ Not available", "⚠️ setExtraHTTPHeaders() only", "⚠️ Limited"),
    ("Cookie Management", "✅ Full context cookies", "⚠️ Via Chrome profile API", "✅ addCookies(), clearCookies()", "✅ Via server"),
    ("localStorage / sessionStorage", "✅ page.evaluate()", "⚠️ Via JavaScript eval", "⚠️ Broken in v3 (storage regression)", "⚠️ Via server"),
    ("iframe / Shadow DOM", "⚠️ Manual frameLocator()", "⚠️ Variable LLM support", "✅ Automatic traversal (v3 44% faster)", "✅ Via server"),
    ("Dialog Handling (alert/confirm)", "✅ page.on('dialog')", "⚠️ Limited", "⚠️ Not well-documented", "⚠️ Limited"),
    ("PDF Rendering", "✅ page.pdf()", "❌ Not available", "❌ Not available", "❌ Not available"),

    # CATEGORY: AI / LLM FEATURES
    ("__cat__", "5. AI / LLM FEATURES"),
    ("LLM Integration Architecture", "❌ None", "⚠️ LangChain BaseChatModel", "✅ Vercel AI SDK (15+ providers)", "✅ Via server (all TS providers)"),
    ("Supported LLM Providers", "❌ N/A", "✅ OpenAI, Anthropic, Google, Ollama, etc.", "✅ 15+: OpenAI, Anthropic, Google, Groq, Ollama, Bedrock, etc.", "✅ Via server (all TS providers)"),
    ("Vision / Screenshot to LLM", "❌ None", "✅ Configurable (auto/true/false)", "✅ CUA mode; screenshot tool in agent", "✅ Via server"),
    ("Structured Data Extraction", "⚠️ Manual DOM parsing", "⚠️ Pydantic schema (LLM output)", "✅ Zod schema + extract() primitive", "✅ Via server (Pydantic schemas)"),
    ("Action Planning (observe)", "❌ None", "⚠️ Implicit in agent loop", "✅ observe() — returns action map", "✅ Via server"),
    ("Single-Step AI Action (act)", "❌ None", "❌ Full loop only", "✅ act() — atomic NL action", "✅ Via server"),
    ("Multi-Step Autonomous Agent", "❌ None", "✅ Core feature", "✅ agent() — DOM/CUA/Hybrid modes", "✅ Via server"),
    ("Computer Use (CUA) Mode", "❌ None", "⚠️ Vision-based but no CUA API", "✅ Claude/GPT-4o/Gemini CUA", "✅ Via server"),
    ("Custom Tools / Actions", "❌ N/A", "✅ @tools.action() decorator", "✅ Agent tools API", "⚠️ Limited via server"),
    ("Dual LLM (reasoning + extraction)", "❌ None", "✅ page_extraction_llm param", "⚠️ Per-call model override", "⚠️ Limited"),
    ("Multi-Turn Agent Continuation", "❌ None", "⚠️ Via history pass-through", "✅ messages param in agent.execute()", "⚠️ Limited"),

    # CATEGORY: PERFORMANCE
    ("__cat__", "6. PERFORMANCE CHARACTERISTICS"),
    ("Deterministic Action Speed", "✅ <100ms per action", "❌ N/A (always LLM-bound)", "✅ ~2-3s (cached) / 20-30s (uncached)", "✅ Via server (similar)"),
    ("AI Action Latency (uncached)", "❌ N/A", "⚠️ 30-120s per task", "⚠️ 2-8s per act() call", "⚠️ Via server + HTTP overhead"),
    ("AI Action Cost (per call)", "❌ N/A", "$0.05–1.00 per task", "$0.01–0.05+ per act() call (uncached)", "Same as TS via server"),
    ("Auto-Caching Benefit", "❌ None", "⚠️ Cloud SDK only", "✅ 10-100x faster on repeats, $0 LLM", "✅ Via server"),
    ("Parallelism (multi-platform)", "✅ asyncio.gather() native", "⚠️ Limited by LLM rate limits", "✅ Promise.all() in TS", "⚠️ asyncio + HTTP calls"),
    ("Flash Mode (skip CoT)", "❌ N/A", "✅ flash_mode=True", "❌ Not available", "❌ Not available"),

    # CATEGORY: PRODUCTION FEATURES
    ("__cat__", "7. PRODUCTION & CLOUD FEATURES"),
    ("Session Persistence (cookies)", "✅ storageState(), profiles", "✅ Chrome profile reuse", "⚠️ Broken in v3 (regression)", "⚠️ Via server (same regression)"),
    ("Multi-Session Concurrency", "✅ Full asyncio concurrency", "⚠️ Limited by browser memory", "✅ Browserbase infra (cloud)", "✅ Via Browserbase cloud"),
    ("Anti-Bot / Stealth Mode", "❌ Default automation fingerprint", "⚠️ Cloud plan only", "⚠️ Browserbase cloud (Scale plan)", "⚠️ Browserbase cloud only"),
    ("CAPTCHA Solving", "❌ None", "⚠️ Cloud plan only", "⚠️ Browserbase cloud (Developer+)", "⚠️ Browserbase cloud only"),
    ("Session Recording / Replay", "❌ None built-in", "⚠️ video extra (OSS)", "✅ Browserbase dashboard (cloud)", "✅ Browserbase dashboard (cloud)"),
    ("AI Decision Observability", "❌ None", "⚠️ history.model_thoughts()", "✅ Stagehand Inspector tab (cloud)", "✅ Via Browserbase cloud"),
    ("Cloud Browser Infrastructure", "❌ DIY", "⚠️ Browser-Use Cloud (separate service)", "✅ Browserbase (tightly integrated)", "✅ Browserbase (tightly integrated)"),
    ("MCP Server Integration", "❌ None", "✅ Full MCP server (OSS+Cloud)", "✅ Browserbase MCP server", "❌ Not available in Python SDK"),

    # CATEGORY: DEVELOPER EXPERIENCE
    ("__cat__", "8. DEVELOPER EXPERIENCE"),
    ("Test Framework Integration", "✅ Designed for testing (pytest)", "⚠️ Not test-focused", "⚠️ Not test-focused", "⚠️ Not test-focused"),
    ("Code Generation / Recorder", "✅ codegen tool included", "❌ None", "❌ None", "❌ None"),
    ("Debugging Tools", "✅ Inspector, trace viewer, slowMo", "⚠️ Screenshots + history", "✅ Stagehand Inspector (cloud)", "⚠️ Server logs + history"),
    ("Type Safety (TypeScript)", "✅ Full typed API (Python+TS)", "⚠️ Python typed (Pydantic)", "✅ Full TypeScript + Zod", "⚠️ Python typed (Pydantic)"),
    ("Documentation Quality", "✅ Excellent (Microsoft-backed)", "⚠️ Good, evolving rapidly", "✅ Good (Browserbase-backed)", "⚠️ Limited (newer SDK)"),
    ("CI/CD Suitability", "✅ Headless, CI-native", "⚠️ Headless but LLM-dependent", "✅ Headless + Browserbase cloud CI", "⚠️ Requires running server"),
    ("API Stability", "✅ Stable (1.x, versioned)", "⚠️ Rapid changes (sub-1.0)", "⚠️ Recent breaking change (v3)", "⚠️ Newer, may change"),
    ("Community Size", "✅ Very large (70k stars, Microsoft)", "✅ Very large (86k stars, fastest-growing)", "⚠️ Growing (22k stars)", "⚠️ Small (Python SDK is new)"),

    # CATEGORY: MULTAI-SPECIFIC FIT
    ("__cat__", "9. MULTAI-SPECIFIC COMPATIBILITY"),
    ("Python-Native (no server needed)", "✅ Fully native Python", "✅ Fully native Python", "❌ TypeScript server required", "❌ HTTP client to TS server"),
    ("Async-First (asyncio)", "✅ playwright.async_api", "✅ Full asyncio", "❌ No native Python asyncio", "⚠️ asyncio HTTP calls only"),
    ("CDP Connect to Running Chrome", "✅ connect_over_cdp()", "✅ BrowserSession(cdp_url=...)", "✅ connectURL() (TS) / CDP bridge", "⚠️ Via server process"),
    ("Per-Step Fallback Pattern", "⚠️ No fallback built-in (MultAI adds this)", "⚠️ Full loop, not per-step", "✅ act()/observe() per step", "✅ Via server (same)"),
    ("Custom Platform Selectors", "✅ Full CSS/XPath control", "⚠️ Selector-like (DOM index)", "✅ Explicit selectors + AI fallback", "✅ Via server"),
    ("Response Text Extraction", "⚠️ Manual DOM parsing (platform-specific)", "⚠️ LLM extraction (non-deterministic)", "✅ extract() + Zod schema", "✅ Via server (Pydantic)"),
    ("Parallel Platform Automation", "✅ asyncio.gather() — 7 tabs native", "⚠️ Sequential or multi-instance", "✅ Promise.all() in TS", "⚠️ asyncio HTTP pool"),
    ("Cost at Scale (7 platforms)", "✅ Zero LLM cost", "⚠️ ~$0.05-0.50/run (7 platforms)", "⚠️ High if uncached ($0.07-0.35/run)", "⚠️ Same as TS + HTTP overhead"),
    ("Zero-Infrastructure Overhead", "✅ No extra process", "✅ No extra process", "❌ Requires stagehand server", "❌ Requires running TS server"),
    ("Existing Code Reuse", "✅ 100% (already in use)", "✅ 100% (already in use)", "❌ Full rewrite required", "❌ Full rewrite required"),
]

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 2 — Pros & Cons for MultAI
# ══════════════════════════════════════════════════════════════════════════════

PROS_CONS = {
    "pros": [
        (
            "Single unified API for hybrid automation",
            "Stagehand's four primitives (act, extract, observe, agent) consolidate what MultAI currently splits across two dependencies (Playwright + browser-use). One framework handles both deterministic and agentic steps.",
            "Medium",
        ),
        (
            "Auto-caching converts AI actions to deterministic scripts",
            "First time a platform's UI is navigated, Stagehand uses AI to find elements and caches the resolved XPath. Subsequent runs replay deterministically — same reliability as Playwright selectors, automatically discovered. For MultAI's 7 stable-ish platforms, most runs would eventually be cache hits.",
            "High",
        ),
        (
            "extract() with Zod/Pydantic schema for structured response extraction",
            "MultAI currently uses custom per-platform extract_response() methods with brittle DOM parsing. Stagehand's extract() with a schema would yield structured, typed output with AI fallback if the DOM changes — eliminating the most maintenance-heavy code.",
            "High",
        ),
        (
            "observe() enables single-LLM-call multi-step planning",
            "Calling observe() once identifies all needed elements, reducing per-step LLM calls from N to 1 for N-step sequences. For 7 platforms in parallel, this could cut AI API costs significantly versus browser-use's per-step loop.",
            "Medium",
        ),
        (
            "15+ LLM providers via Vercel AI SDK",
            "Stagehand supports Anthropic, OpenAI, Google, Groq, Ollama, Bedrock, and others — more than browser-use's LangChain adapter list. MultAI already uses Anthropic + Google; switching between them per step or for cost optimisation would be trivial.",
            "Low",
        ),
        (
            "Browserbase cloud enables session replay and AI decision observability",
            "The Stagehand Inspector tab shows exactly which elements the AI chose and why — a debugging capability MultAI currently lacks entirely. Useful for diagnosing why a platform automation fails.",
            "Medium",
        ),
        (
            "CDP-native v3 architecture (44% faster for iframe/shadow DOM)",
            "If MultAI were in TypeScript, v3's CDP-native approach would outperform Playwright's actionability-check overhead. Less relevant for current Python stack but worth noting for any future TypeScript rewrite.",
            "Low",
        ),
        (
            "Active development and growing ecosystem",
            "21,800+ stars, multiple releases/month, Browserbase backing. Less risk of abandonment than browser-use's independent maintainer model. MCP server, CrewAI, LangChain integrations already available.",
            "Low",
        ),
    ],
    "cons": [
        (
            "TypeScript-primary: Python SDK is a thin HTTP client, not a native library",
            "MultAI is 100% Python. The Stagehand Python SDK is not a port — it is a Python HTTP client that sends requests to a separately-running TypeScript Stagehand server process. This means: (1) an additional server must be started alongside MultAI, (2) the Python asyncio integration is shallow (HTTP calls), (3) Python-specific features like direct Page object manipulation are lost. This alone makes Stagehand a poor fit for MultAI's current architecture.",
            "Critical",
        ),
        (
            "Full rewrite required — zero code reuse",
            "BasePlatform, all 7 platform drivers, agent_fallback.py, and the orchestrator engine are all Python. Adopting Stagehand (even via Python SDK) requires rewriting the entire automation layer from scratch. MultAI's existing code works and is well-tested; the rewrite risk is not justified by the benefits.",
            "Critical",
        ),
        (
            "Requires running a separate Stagehand TypeScript server process",
            "The Python SDK requires a running stagehand-server (Node.js/TypeScript). MultAI's deployment would gain a new runtime dependency (Node.js), a new process to manage, and a new failure point. This complicates setup.sh, Docker images, and CI.",
            "High",
        ),
        (
            "Session storage state regression in v3",
            "storageState() and userDataDir/preserveUserDataDir are broken in Stagehand v3. Cookie persistence across sessions does not work as expected. MultAI relies on Chrome profiles staying authenticated — this regression is a blocker.",
            "High",
        ),
        (
            "Higher LLM cost per run (uncached first-run economics)",
            "On first run (cache miss), each act() call costs $0.01–0.05 in LLM fees. MultAI runs 7 platforms per query. Even with quick caching, if MultAI handles variable prompts (which changes cache keys), many runs will be cache misses. Browser-use's cost is similar, but Stagehand's cost advantage appears only after the cache warms up for each platform/URL/DOM combination.",
            "High",
        ),
        (
            "CUA mode limitations (no variables, streaming, or abort signals)",
            "Computer Use Agent mode — the closest equivalent to browser-use's vision-based approach — does not support: variable substitution (security concern for passwords), streaming, callbacks, abort signals, or message continuation. These are all features MultAI's BasePlatform currently relies on for clean handoffs.",
            "Medium",
        ),
        (
            "Chromium-only (CDP constraint)",
            "Stagehand v3 is CDP-native and supports only Chrome/Chromium. MultAI already uses Chrome exclusively, so this is not a blocking limitation — but it removes optionality for running against non-Chrome browsers in the future.",
            "Low",
        ),
        (
            "Python SDK lacks MCP client support",
            "Stagehand's MCP client functionality (for integrating external MCP servers into agent workflows) is unavailable in the Python SDK. If MultAI wanted to leverage MCP in browser agents, it would be stuck waiting for parity.",
            "Low",
        ),
        (
            "API instability: v3 was a breaking change less than 3 months ago",
            "Stagehand v3 (January 2025) removed the Playwright dependency, breaking all v2.x code. MultAI has already had to manage browser-use breaking changes across 120+ releases. Adding another rapidly-changing dependency increases maintenance burden.",
            "Medium",
        ),
        (
            "Benchmarked at ~85% per atomic action / ~65% end-to-end accuracy",
            "MultAI's Playwright automation, when working, achieves near-100% accuracy on the deterministic paths. Stagehand's AI primitives introduce a 15% per-action failure rate on uncached operations. For 7 platforms × multiple steps, compounding errors mean multi-step workflows could fail more often than pure Playwright.",
            "Medium",
        ),
    ],
}


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 3 — Executive Summary / Recommendation
# ══════════════════════════════════════════════════════════════════════════════

SUMMARY_ROWS = [
    ("MULTAI CONTEXT", "", "", ""),
    ("Automation Stack", "Current", "Playwright 1.58 (deterministic primary) + Browser-Use 0.12.2 (agentic fallback)", ""),
    ("Architecture Pattern", "Current", "BasePlatform: Playwright first → browser-use fallback on selector failure", ""),
    ("Language", "Current", "Python (100%) — asyncio-first", ""),
    ("Platforms Automated", "Current", "7 (Claude, ChatGPT, Copilot, Perplexity, Grok, DeepSeek, Gemini)", ""),
    ("", "", "", ""),
    ("STAGEHAND OVERVIEW", "", "", ""),
    ("Framework", "Stagehand", "MIT, Browserbase-backed, v3.2.0 (March 2025)", ""),
    ("Architecture", "Stagehand", "Hybrid deterministic+agentic. act()/extract()/observe()/agent() primitives. Auto-caching converts AI-discovered workflows to deterministic scripts.", ""),
    ("Python Support", "Stagehand", "HTTP client only — requires separately-running TypeScript server (Node.js process)", ""),
    ("Unique Strength", "Stagehand", "Auto-caching (10-100x speedup on repeats), structured extract(), 15+ LLM providers, Browserbase cloud integration", ""),
    ("", "", "", ""),
    ("VERDICT", "", "", ""),
    ("Recommendation", "⚠️ NOT RECOMMENDED for MultAI at this time", "", ""),
    ("Primary Blocker", "Python SDK is a thin HTTP client — not a native Python library. Full rewrite of all automation code required with no net capability gain for MultAI's use case.", "", ""),
    ("Secondary Blockers", "Session storage regression (v3), separate Node.js server process required, zero code reuse from current ~800-line Playwright automation layer.", "", ""),
    ("", "", "", ""),
    ("ALTERNATIVE RECOMMENDATIONS", "", "", ""),
    ("Rec 1", "Keep current stack", "Playwright + browser-use hybrid already implements the same pattern Stagehand promises. No migration needed.", ""),
    ("Rec 2", "Selectively adopt Stagehand concepts", "Implement auto-caching logic natively in Python (cache resolved XPath selectors after first AI-discovered path). This gives the primary Stagehand benefit without adopting the framework.", ""),
    ("Rec 3", "Re-evaluate when Python SDK matures", "Stagehand Python SDK is new (May 2025). Monitor for direct Playwright integration, native asyncio support, and feature parity with the TypeScript SDK before reconsidering.", ""),
    ("Rec 4", "Upgrade browser-use to 0.12.6", "browser-use 0.12.6 (April 2, 2026) is current vs MultAI's pinned 0.12.2. Minor upgrade with bug fixes and ChatBrowserUse-2 model support improving agent accuracy.", ""),
]


def build_matrix_sheet(wb):
    ws = wb.create_sheet("Comparison Matrix")
    ws.freeze_panes = "B3"
    ws.sheet_view.showGridLines = False

    # Title
    ws.merge_cells("A1:E1")
    title = ws["A1"]
    title.value = "Stagehand vs Playwright vs Browser-Use — Capability & Feature Comparison Matrix"
    title.font = Font(name="Arial", bold=True, size=14, color=C_HEADER_FG)
    title.fill = hfill(C_HEADER_BG)
    title.alignment = centered(wrap=False)
    ws.row_dimensions[1].height = 28

    # Column headers
    write_header_row(ws, 2, MATRIX_COLS, sizes=[10, 10, 10, 10, 10])
    ws.row_dimensions[2].height = 42

    # Column widths
    set_col_width(ws, 1, 38)
    for c in range(2, 6):
        set_col_width(ws, c, 32)

    row = 3
    alt = False
    for entry in MATRIX_DATA:
        if entry[0] == "__cat__":
            write_category(ws, row, entry[1], 5)
            row += 1
            alt = False
        else:
            write_data_row(ws, row, list(entry), alt=alt)
            row += 1
            alt = not alt

    # Legend row
    row += 1
    legend_cell = ws.cell(row=row, column=1, value="Legend:   ✅ Full support   ⚠️ Partial / conditional   ❌ Not available / not applicable")
    legend_cell.font = Font(name="Arial", italic=True, size=8, color="595959")
    legend_cell.alignment = left_align(wrap=False)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    ws.row_dimensions[row].height = 16


def build_pros_cons_sheet(wb):
    ws = wb.create_sheet("Pros & Cons (MultAI)")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"

    # Title
    ws.merge_cells("A1:D1")
    t = ws["A1"]
    t.value = "Pros & Cons: Replacing Playwright + Browser-Use with Stagehand in MultAI"
    t.font = Font(name="Arial", bold=True, size=14, color=C_HEADER_FG)
    t.fill = hfill(C_HEADER_BG)
    t.alignment = centered()
    ws.row_dimensions[1].height = 28

    headers = ["#", "Pro / Con", "Detail", "Impact"]
    col_widths = [4, 32, 65, 12]
    write_header_row(ws, 2, headers)
    ws.row_dimensions[2].height = 20
    for i, w in enumerate(col_widths, 1):
        set_col_width(ws, i, w)

    row = 3
    # Pros section header
    ws.merge_cells(f"A{row}:D{row}")
    ch = ws[f"A{row}"]
    ch.value = "PROS — Arguments FOR adopting Stagehand"
    ch.fill = hfill("375623")
    ch.font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    ch.alignment = centered()
    ws.row_dimensions[row].height = 18
    row += 1

    for i, (title, detail, impact) in enumerate(PROS_CONS["pros"], 1):
        alt = (i % 2 == 0)
        bg = "EAF4E2" if not alt else "D9EAD3"
        for c_idx, val in enumerate([str(i), title, detail, impact], 1):
            cell = ws.cell(row=row, column=c_idx, value=val)
            cell.fill = hfill(bg)
            cell.border = thin_border()
            cell.alignment = left_align(wrap=True)
            if c_idx == 1:
                cell.font = Font(name="Arial", bold=True, size=9, color=C_YES)
                cell.alignment = centered()
            elif c_idx == 2:
                cell.font = Font(name="Arial", bold=True, size=9, color=C_YES)
            elif c_idx == 4:
                imp_color = {"High": C_YES, "Medium": C_PARTIAL, "Low": "595959"}.get(val, "000000")
                cell.font = Font(name="Arial", bold=True, size=9, color=imp_color)
                cell.alignment = centered()
            else:
                cell.font = Font(name="Arial", size=9)
        ws.row_dimensions[row].height = 55
        row += 1

    row += 1
    # Cons section header
    ws.merge_cells(f"A{row}:D{row}")
    cc = ws[f"A{row}"]
    cc.value = "CONS — Arguments AGAINST adopting Stagehand"
    cc.fill = hfill("9C0006")
    cc.font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    cc.alignment = centered()
    ws.row_dimensions[row].height = 18
    row += 1

    for i, (title, detail, impact) in enumerate(PROS_CONS["cons"], 1):
        alt = (i % 2 == 0)
        bg = C_RED_BG if not alt else "FAD7C3"
        for c_idx, val in enumerate([str(i), title, detail, impact], 1):
            cell = ws.cell(row=row, column=c_idx, value=val)
            cell.fill = hfill(bg)
            cell.border = thin_border()
            cell.alignment = left_align(wrap=True)
            if c_idx == 1:
                cell.font = Font(name="Arial", bold=True, size=9, color=C_NO)
                cell.alignment = centered()
            elif c_idx == 2:
                cell.font = Font(name="Arial", bold=True, size=9, color=C_NO)
            elif c_idx == 4:
                imp_color = {
                    "Critical": C_NO, "High": C_PARTIAL,
                    "Medium": "7D4900", "Low": "595959"
                }.get(val, "000000")
                cell.font = Font(name="Arial", bold=True, size=9, color=imp_color)
                cell.alignment = centered()
            else:
                cell.font = Font(name="Arial", size=9)
        ws.row_dimensions[row].height = 55
        row += 1


def build_summary_sheet(wb):
    ws = wb.create_sheet("Executive Summary")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:D1")
    t = ws["A1"]
    t.value = "Executive Summary & Recommendation — Stagehand for MultAI"
    t.font = Font(name="Arial", bold=True, size=14, color=C_HEADER_FG)
    t.fill = hfill(C_HEADER_BG)
    t.alignment = centered()
    ws.row_dimensions[1].height = 28

    set_col_width(ws, 1, 26)
    set_col_width(ws, 2, 14)
    set_col_width(ws, 3, 75)
    set_col_width(ws, 4, 12)

    write_header_row(ws, 2, ["Section", "Attribute", "Details", ""], sizes=[10, 10, 10, 10])
    ws.row_dimensions[2].height = 20

    section_colors = {
        "MULTAI CONTEXT": "1F4E79",
        "STAGEHAND OVERVIEW": "375623",
        "VERDICT": "9C0006",
        "ALTERNATIVE RECOMMENDATIONS": "7D4900",
    }

    row = 3
    current_section = ""
    alt = False
    for entry in SUMMARY_ROWS:
        if not any(entry):
            row += 1
            alt = False
            continue
        section, attr, detail, _ = entry
        if section in section_colors:
            current_section = section
            ws.merge_cells(f"A{row}:D{row}")
            sc = ws[f"A{row}"]
            sc.value = section
            sc.fill = hfill(section_colors[section])
            sc.font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
            sc.alignment = centered()
            ws.row_dimensions[row].height = 22
            row += 1
            alt = False
            continue

        bg = C_ALT if alt else C_WHITE
        for c_idx, val in enumerate([section, attr, detail, ""], 1):
            cell = ws.cell(row=row, column=c_idx, value=val)
            cell.border = thin_border()
            cell.alignment = left_align(wrap=True)

            if c_idx == 3 and "⚠️ NOT RECOMMENDED" in str(val):
                cell.fill = hfill(C_RED_BG)
                cell.font = Font(name="Arial", bold=True, size=10, color=C_NO)
            elif c_idx == 3 and str(val).startswith("Playwright +"):
                cell.fill = hfill(C_GREEN_BG)
                cell.font = Font(name="Arial", bold=True, size=9, color=C_YES)
            else:
                cell.fill = hfill(bg)
                cell.font = Font(name="Arial", size=9)

        ws.row_dimensions[row].height = 40
        row += 1
        alt = not alt


def main():
    import os
    os.makedirs("reports", exist_ok=True)

    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    build_summary_sheet(wb)
    build_matrix_sheet(wb)
    build_pros_cons_sheet(wb)

    wb.save(OUTPUT)
    print(f"Saved: {OUTPUT}")


if __name__ == "__main__":
    main()
