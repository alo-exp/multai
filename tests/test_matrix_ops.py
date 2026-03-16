"""Unit tests for matrix_ops module.

Tests UT-MX-01 through UT-MX-09.

Uses subprocess to invoke matrix_ops.py as a CLI to avoid import complexity.
"""

import json
import subprocess
import sys
import tempfile
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter

PROJECT_ROOT = Path(__file__).parent.parent
MATRIX_OPS = PROJECT_ROOT / "skills" / "comparator" / "matrix_ops.py"

TICK = "\u2714"  # checkmark


def _create_test_matrix(path: str, with_title: bool = True) -> None:
    """Create a minimal test XLSX matrix.

    Layout (with_title=True):
      Row 1: Title (merged A1:E1)
      Row 2: Headers — "", "Priority", "PlatformA", "PlatformB"
      Row 3: COUNTIF formulas
      Row 4: Score formulas
      Row 5: Category row "Core Features" (merged)
      Row 6: Feature "CI/CD Pipelines", "Critical", tick, tick
      Row 7: Feature "Container Support", "High", tick, ""
      Row 8: Feature "Monitoring", "Medium", "", tick
      Row 9: Category row "Advanced Features" (merged)
      Row 10: Feature "AI Ops", "Low", tick, ""
      Row 11: Feature "GitOps", "Very High", "", tick
    """
    wb = openpyxl.Workbook()
    ws = wb.active

    if with_title:
        ws.cell(1, 1).value = "Test Comparison Matrix"
        ws.merge_cells("A1:E1")
        header_row = 2
        total_row = 3
        score_row = 4
        data_start = 5
    else:
        header_row = 1
        total_row = 2
        score_row = 3
        data_start = 4

    # Headers
    ws.cell(header_row, 1).value = ""
    ws.cell(header_row, 2).value = "Priority"
    ws.cell(header_row, 3).value = "PlatformA"
    ws.cell(header_row, 4).value = "PlatformB"

    last_col = 4

    # COUNTIF and score rows (placeholder formulas)
    for c in range(3, last_col + 1):
        ltr = get_column_letter(c)
        ws.cell(total_row, c).value = f'=COUNTIF({ltr}{data_start}:{ltr}1048576,"?*")'
        ws.cell(score_row, c).value = 0  # Placeholder

    # Category 1
    r = data_start
    ws.cell(r, 1).value = "Core Features"
    ws.merge_cells(f"A{r}:{get_column_letter(last_col)}{r}")

    # Feature rows
    features = [
        ("CI/CD Pipelines", "Critical", True, True),
        ("Container Support", "High", True, False),
        ("Monitoring", "Medium", False, True),
    ]
    for feat, prio, tick_a, tick_b in features:
        r += 1
        ws.cell(r, 1).value = feat
        ws.cell(r, 2).value = prio
        ws.cell(r, 3).value = TICK if tick_a else None
        ws.cell(r, 4).value = TICK if tick_b else None

    # Category 2
    r += 1
    ws.cell(r, 1).value = "Advanced Features"
    ws.merge_cells(f"A{r}:{get_column_letter(last_col)}{r}")

    features2 = [
        ("AI Ops", "Low", True, False),
        ("GitOps", "Very High", False, True),
    ]
    for feat, prio, tick_a, tick_b in features2:
        r += 1
        ws.cell(r, 1).value = feat
        ws.cell(r, 2).value = prio
        ws.cell(r, 3).value = TICK if tick_a else None
        ws.cell(r, 4).value = TICK if tick_b else None

    wb.save(path)


def _run_matrix_ops(*args: str) -> dict:
    """Run matrix_ops.py as subprocess and parse JSON output."""
    cmd = [sys.executable, str(MATRIX_OPS)] + list(args)
    result = subprocess.run(
        cmd,
        capture_output=True, text=True,
        cwd=str(PROJECT_ROOT),
    )
    assert result.returncode == 0, (
        f"matrix_ops.py failed:\nstdout: {result.stdout}\nstderr: {result.stderr}"
    )
    return json.loads(result.stdout)


class TestMatrixInfo:
    """Tests for the 'info' command."""

    def test_ut_mx_01_info_returns_platforms(self):
        """UT-MX-01: info command returns platform names."""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name
        _create_test_matrix(path)
        data = _run_matrix_ops("info", "--src", path)
        assert "PlatformA" in data["platforms"]
        assert "PlatformB" in data["platforms"]
        assert data["platform_count"] == 2

    def test_ut_mx_02_info_returns_categories(self):
        """UT-MX-02: info returns correct category names."""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name
        _create_test_matrix(path)
        data = _run_matrix_ops("info", "--src", path)
        assert "Core Features" in data["categories"]
        assert "Advanced Features" in data["categories"]
        assert data["category_count"] == 2

    def test_ut_mx_03_info_total_features(self):
        """UT-MX-03: info returns correct total feature count."""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name
        _create_test_matrix(path)
        data = _run_matrix_ops("info", "--src", path)
        assert data["total_features"] == 5

    def test_ut_mx_04_info_detects_with_title_layout(self):
        """UT-MX-04: info detects 'with_title' layout."""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name
        _create_test_matrix(path, with_title=True)
        data = _run_matrix_ops("info", "--src", path)
        assert data["layout"] == "with_title"

    def test_ut_mx_05_info_detects_no_title_layout(self):
        """UT-MX-05: info detects 'no_title' layout."""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name
        _create_test_matrix(path, with_title=False)
        data = _run_matrix_ops("info", "--src", path)
        assert data["layout"] == "no_title"


class TestMatrixScores:
    """Tests for the 'scores' command."""

    def test_ut_mx_06_scores_returns_rankings(self):
        """UT-MX-06: scores command returns ranked platforms."""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name
        _create_test_matrix(path)
        data = _run_matrix_ops("scores", "--src", path)
        rankings = data["rankings"]
        assert len(rankings) == 2
        names = [r["platform"] for r in rankings]
        assert "PlatformA" in names
        assert "PlatformB" in names


class TestMatrixExtractFeatures:
    """Tests for the 'extract-features' command."""

    def test_ut_mx_07_extract_features_groups_by_category(self):
        """UT-MX-07: extract-features returns features grouped by category."""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name
        _create_test_matrix(path)
        data = _run_matrix_ops("extract-features", "--src", path)
        assert data["total_categories"] == 2
        assert data["total_features"] == 5
        cat_names = [c["name"] for c in data["categories"]]
        assert "Core Features" in cat_names


class TestMatrixVerify:
    """Tests for the 'verify' command."""

    def test_ut_mx_08_verify_returns_tick_lists(self):
        """UT-MX-08: verify returns per-platform tick lists."""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name
        _create_test_matrix(path)
        data = _run_matrix_ops("verify", "--src", path)
        assert "PlatformA" in data
        assert "CI/CD Pipelines" in data["PlatformA"]["ticked"]
        assert data["PlatformA"]["tick_count"] == 3  # CI/CD, Container, AI Ops


class TestMatrixAddPlatform:
    """Tests for the 'add-platform' command."""

    def test_ut_mx_09_add_platform(self):
        """UT-MX-09: add-platform adds a new column."""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            src_path = f.name
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            out_path = f.name

        _create_test_matrix(src_path)

        # Create features JSON
        features = {"CI/CD Pipelines": True, "Container Support": True, "Monitoring": False, "AI Ops": False, "GitOps": True}
        with tempfile.NamedTemporaryFile(mode="w", suffix=".json", delete=False) as f:
            json.dump(features, f)
            feat_path = f.name

        data = _run_matrix_ops(
            "add-platform",
            "--src", src_path,
            "--out", out_path,
            "--platform", "PlatformC",
            "--features", feat_path,
        )
        assert data["ticks_applied"] == 3  # CI/CD, Container, GitOps
        assert data["platform_col"] == 5  # New column after D

        # Verify the output file
        info_data = _run_matrix_ops("info", "--src", out_path)
        assert "PlatformC" in info_data["platforms"]
        assert info_data["platform_count"] == 3
